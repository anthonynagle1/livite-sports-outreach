"""Aggregate catering revenue by month and platform.

Primary source: Excel Tracker sheet in New Profit Calc (1).xlsx
  - Forkable (col 10), EZ Cater (col 11), Cater2Me (col 12): Nov 2024+
  - Catering Delivery (col 17): Toast catering orders, Jan 2026+
  - Catering Fees (col 35): commission paid to platforms

Secondary source: Toast POS cached data (.tmp/<YYYYMMDD>/OrderDetails.csv)
  - "Catering Delivery" dining option -- supplements Excel for newer dates
  - For Toast Catering: per month, take max(Excel total, POS total) to avoid
    double-counting while capturing the most complete data.

Dashboard source: Washington Square catering CSV for Toast revenue + customer enrichment.
  - Same orders as Toast POS -- per month, take max(POS, WS CSV) to capture
    most complete data (POS cache may be partial).
  - Also used for customer names, companies, delivery addresses, guest counts.
"""

from __future__ import annotations

import csv
import logging
import os
import time
from collections import defaultdict
from datetime import datetime

import pandas as pd

logger = logging.getLogger(__name__)

_BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
_EXCEL_PATH = os.path.join(_BASE_DIR, "data", "New Profit Calc (1).xlsx")
_WS_CSV_PATH = os.path.join(_BASE_DIR, "data", "livite-washington-square_orders.csv")
_TMP_DIR = os.path.join(_BASE_DIR, ".tmp")

# Cache with TTL
_cache = {"ts": 0, "data": None}
_CACHE_TTL = 3600  # 1 hour

# Dashboard cache (separate from monthly cache)
_dash_cache = {"ts": 0, "data": None}


def _parse_tracker_catering():
    """Parse catering data from the Tracker sheet.

    Returns tuple: (orders, toast_catering_monthly)
      orders: list of {date, platform, subtotal} for Forkable/EZCater/Cater2Me/fees
      toast_catering_monthly: {month: total} from Excel's Catering Delivery column
    """
    if not os.path.exists(_EXCEL_PATH):
        logger.warning("Excel tracker not found: %s", _EXCEL_PATH)
        return [], {}

    try:
        from openpyxl import load_workbook
        wb = load_workbook(_EXCEL_PATH, read_only=True, data_only=True)
    except Exception as e:
        logger.error("Failed to open Excel tracker: %s", e)
        return [], {}

    if "Tracker" not in wb.sheetnames:
        logger.warning("'Tracker' sheet not found in Excel file")
        wb.close()
        return [], {}

    ws = wb["Tracker"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return [], {}

    # Build column index -- first occurrence wins (duplicate "Date" at cols 0 and 48)
    header = [str(c).strip() if c else "" for c in rows[0]]
    col_map = {}
    for i, h in enumerate(header):
        if h and h not in col_map:
            col_map[h] = i

    date_idx = col_map.get("Date")       # col 0
    fork_idx = col_map.get("Forkable")    # col 10
    ez_idx = col_map.get("EZ Cater")      # col 11
    c2m_idx = col_map.get("Cater2Me")     # col 12
    cat_idx = col_map.get("Catering Delivery")  # col 17
    fees_idx = col_map.get("Catering Fees")     # col 35

    if date_idx is None:
        logger.warning("'Date' column not found in Tracker sheet")
        return [], {}

    def _num(row, idx):
        if idx is None or idx >= len(row):
            return 0.0
        v = row[idx]
        if v is None:
            return 0.0
        try:
            return float(v)
        except (ValueError, TypeError):
            return 0.0

    orders = []
    toast_monthly = defaultdict(float)

    for row in rows[1:]:
        if not row or len(row) <= date_idx:
            continue

        raw_date = row[date_idx]
        if raw_date is None:
            continue

        # Parse date -- skip footer/summary rows
        if isinstance(raw_date, datetime):
            date_str = raw_date.strftime("%Y-%m-%d")
        else:
            continue  # non-datetime values are summary rows

        month = date_str[:7]

        fork_val = _num(row, fork_idx)
        ez_val = _num(row, ez_idx)
        c2m_val = _num(row, c2m_idx)
        cat_val = _num(row, cat_idx)
        fees_val = _num(row, fees_idx)

        if fork_val > 0:
            orders.append({"date": date_str, "platform": "Forkable", "subtotal": fork_val})
        if ez_val > 0:
            orders.append({"date": date_str, "platform": "EZCater", "subtotal": ez_val})
        if c2m_val > 0:
            orders.append({"date": date_str, "platform": "Cater2me", "subtotal": c2m_val})
        if fees_val > 0:
            orders.append({"date": date_str, "platform": "_fees", "subtotal": fees_val})
        if cat_val > 0:
            toast_monthly[month] += cat_val

    return orders, dict(toast_monthly)


def _scan_toast_catering():
    """Scan Toast POS cache for catering orders.

    Two detection methods (per day, take the higher):
      1. OrderDetails: "Catering" in Dining Options (Jan 2026+)
      2. ItemSelectionDetails: "Catering" in Menu column (Feb 2025+)

    Returns dict: {month: total_revenue}.
    """
    if not os.path.isdir(_TMP_DIR):
        return {}

    daily = {}  # date_str -> max revenue

    for folder in os.listdir(_TMP_DIR):
        if not (folder.isdigit() and len(folder) == 8):
            continue

        date_str = folder[:4] + "-" + folder[4:6] + "-" + folder[6:8]
        rev_dining = 0.0
        rev_menu = 0.0

        # Method 1: OrderDetails -- "Catering" dining option
        order_file = os.path.join(_TMP_DIR, folder, "OrderDetails.csv")
        if os.path.exists(order_file):
            try:
                df = pd.read_csv(order_file)
                if "Voided" in df.columns:
                    df = df[df["Voided"] == False]
                if "Dining Options" in df.columns and "Amount" in df.columns:
                    cat = df[df["Dining Options"].astype(str).str.contains(
                        "Catering", case=False, na=False
                    )]
                    if len(cat) > 0:
                        rev_dining = cat["Amount"].sum()
            except Exception as e:
                logger.warning("Catering OrderDetails parse failed for %s: %s", folder, e)

        # Method 2: ItemSelectionDetails -- "Catering" menu
        item_file = os.path.join(_TMP_DIR, folder, "ItemSelectionDetails.csv")
        if os.path.exists(item_file):
            try:
                df = pd.read_csv(item_file)
                if "Menu" in df.columns and "Net Price" in df.columns:
                    cat = df[df["Menu"].astype(str).str.contains(
                        "Catering", case=False, na=False
                    )]
                    if len(cat) > 0:
                        rev_menu = cat["Net Price"].sum()
            except Exception as e:
                logger.warning("Catering ItemSelectionDetails parse failed for %s: %s", folder, e)

        # Take the higher of the two methods (avoids double-counting)
        best = max(rev_dining, rev_menu)
        if best > 0:
            daily[date_str] = round(best, 2)

    # Aggregate to monthly
    monthly = defaultdict(float)
    for date_str, rev in daily.items():
        month = date_str[:7]
        monthly[month] += rev

    return dict(monthly)


def _scan_toast_catering_detail():
    """Scan Toast POS cache for detailed catering data.

    Returns:
        daily_orders: list of {date, revenue, order_count}
        top_items: list of {item, qty, revenue} from ItemSelectionDetails
    """
    if not os.path.isdir(_TMP_DIR):
        return [], []

    daily_orders = []
    item_agg = defaultdict(lambda: {"qty": 0, "revenue": 0.0})

    for folder in sorted(os.listdir(_TMP_DIR)):
        if not (folder.isdigit() and len(folder) == 8):
            continue

        date_str = folder[:4] + "-" + folder[4:6] + "-" + folder[6:8]

        # Order-level from OrderDetails
        order_file = os.path.join(_TMP_DIR, folder, "OrderDetails.csv")
        if os.path.exists(order_file):
            try:
                df = pd.read_csv(order_file)
                if "Voided" in df.columns:
                    df = df[df["Voided"] == False]
                if "Dining Options" in df.columns and "Amount" in df.columns:
                    cat = df[df["Dining Options"].astype(str).str.contains(
                        "Catering", case=False, na=False
                    )]
                    if len(cat) > 0:
                        daily_orders.append({
                            "date": date_str,
                            "revenue": round(cat["Amount"].sum(), 2),
                            "order_count": len(cat),
                        })
            except Exception as e:
                logger.warning("Catering detail OrderDetails parse failed for %s: %s", folder, e)

        # Item-level from ItemSelectionDetails
        item_file = os.path.join(_TMP_DIR, folder, "ItemSelectionDetails.csv")
        if os.path.exists(item_file):
            try:
                df = pd.read_csv(item_file)
                if "Menu" in df.columns and "Net Price" in df.columns:
                    cat = df[df["Menu"].astype(str).str.contains(
                        "Catering", case=False, na=False
                    )]
                    if len(cat) > 0:
                        name_col = "Menu Item" if "Menu Item" in df.columns else None
                        qty_col = "Qty" if "Qty" in df.columns else None
                        for _, row in cat.iterrows():
                            name = str(row[name_col]) if name_col else "Unknown"
                            qty = int(row[qty_col]) if qty_col and pd.notna(row[qty_col]) else 1
                            price = float(row["Net Price"]) if pd.notna(row["Net Price"]) else 0
                            item_agg[name]["qty"] += qty
                            item_agg[name]["revenue"] += price
            except Exception as e:
                logger.warning("Catering detail ItemSelectionDetails parse failed for %s: %s", folder, e)

    # Build top items list sorted by revenue
    top_items = [
        {"item": name, "qty": d["qty"], "revenue": round(d["revenue"], 2)}
        for name, d in item_agg.items()
    ]
    top_items.sort(key=lambda x: x["revenue"], reverse=True)

    return daily_orders, top_items[:20]


def _parse_ws_csv():
    """Parse Washington Square catering CSV.

    Returns list of order dicts with customer/company data and subtotals.
    Used for customer enrichment and as Toast revenue gap-filler when POS
    cache has no data for a given month (same orders, never double-counted).
    """
    if not os.path.exists(_WS_CSV_PATH):
        logger.info("Washington Square CSV not found: %s", _WS_CSV_PATH)
        return []

    orders = []
    try:
        with open(_WS_CSV_PATH, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                status = (row.get("status") or "").strip().upper()
                voided = (row.get("is_voided") or "").strip().lower()
                if voided == "true":
                    continue
                if status not in ("CLOSED", "CONFIRMED"):
                    continue

                # Parse date: "2/21/25, 12:15 PM"
                raw_date = (row.get("date") or "").strip()
                if not raw_date:
                    continue
                try:
                    dt = datetime.strptime(raw_date, "%m/%d/%y, %I:%M %p")
                    date_str = dt.strftime("%Y-%m-%d")
                except ValueError:
                    continue

                # Parse subtotal
                try:
                    subtotal = float(row.get("subtotal") or 0)
                except (ValueError, TypeError):
                    subtotal = 0.0

                try:
                    tip = float(row.get("tip") or 0)
                except (ValueError, TypeError):
                    tip = 0.0

                try:
                    svc = float(row.get("total_service_charges") or 0)
                except (ValueError, TypeError):
                    svc = 0.0

                try:
                    discount = float(row.get("total_discount") or 0)
                except (ValueError, TypeError):
                    discount = 0.0

                try:
                    guest_count = int(row.get("Guest count") or row.get("number_of_guests") or 0)
                except (ValueError, TypeError):
                    guest_count = 0

                first = (row.get("customer_first_name") or "").strip()
                last = (row.get("customer_last_name") or "").strip()
                customer = (first + " " + last).strip() if first or last else ""
                company = (row.get("company_name") or "").strip()
                event_name = (row.get("event_name") or "").strip()
                event_type = (row.get("event_type") or "").strip()
                created_from = (row.get("created_from") or "").strip()

                addr = (row.get("delivery_info_address_1") or "").strip()
                city = (row.get("delivery_info_city") or "").strip()
                state = (row.get("delivery_info_state") or "").strip()

                orders.append({
                    "date": date_str,
                    "customer": customer,
                    "company": company,
                    "event_name": event_name,
                    "event_type": event_type,
                    "created_from": created_from,
                    "subtotal": subtotal,
                    "tip": tip,
                    "service_charges": svc,
                    "discount": discount,
                    "guest_count": guest_count,
                    "address": addr,
                    "city": city,
                    "state": state,
                    "platform": "Toast",
                })
    except Exception as e:
        logger.error("Failed to parse Washington Square CSV: %s", e)
        return []

    return orders


def _fetch_notion_3p():
    """Fetch non-Toast catering orders from Notion.

    Returns list of {date, platform, subtotal, name}.
    """
    try:
        from .notion import _fetch_all_orders, NON_TOAST_PLATFORMS
    except ImportError:
        logger.warning("Notion catering module not available")
        return []

    all_orders = _fetch_all_orders()
    return [
        o for o in all_orders
        if o.get("platform") in NON_TOAST_PLATFORMS
    ]


def get_catering_dashboard_data(start_month=None, end_month=None):
    """Get all catering data for the dedicated catering dashboard.

    Combines:
      - Toast POS (revenue + item detail) -- primary source for Toast catering
      - Washington Square CSV -- Toast revenue gap-fill when POS cache is empty
      - Notion (non-Toast 3P: Forkable, EZCater, Cater2me) -- additive revenue
      - Excel Tracker (historical 3P fallback)

    Args:
        start_month: "YYYY-MM" to filter from (inclusive), or None for all
        end_month: "YYYY-MM" to filter to (inclusive), or None for all

    Returns dict with keys: monthly, kpis, top_items, customers,
        recent_orders, dow_analysis, order_sources, period
    """
    now = time.time()
    if _dash_cache["data"] is not None and (now - _dash_cache["ts"]) < _CACHE_TTL:
        cached = _dash_cache["data"]
        # Re-filter if period changed
        cached_start = cached.get("period", {}).get("_start_month")
        cached_end = cached.get("period", {}).get("_end_month")
        if cached_start == start_month and cached_end == end_month:
            return cached

    # ── 1. Toast POS: monthly aggregates + detail ──
    pos_toast_monthly = _scan_toast_catering()
    daily_orders, top_items = _scan_toast_catering_detail()

    # ── 2. Toast POS: order count per day for accurate counts ──
    toast_order_counts_monthly = defaultdict(int)
    toast_revenue_monthly = defaultdict(float)
    for d in daily_orders:
        m = d["date"][:7]
        toast_order_counts_monthly[m] += d["order_count"]
        toast_revenue_monthly[m] += d["revenue"]

    # Use max of scan methods for revenue (existing logic)
    for m, rev in pos_toast_monthly.items():
        toast_revenue_monthly[m] = max(toast_revenue_monthly.get(m, 0), rev)

    # ── 3. Excel Tracker: 3P platforms + fees ──
    tracker_orders, excel_toast = _parse_tracker_catering()

    excel_3p_monthly = defaultdict(lambda: defaultdict(float))
    excel_3p_counts = defaultdict(lambda: defaultdict(int))
    fees_monthly = defaultdict(float)

    for order in tracker_orders:
        m = order["date"][:7]
        plat = order["platform"]
        if plat == "_fees":
            fees_monthly[m] += order["subtotal"]
        else:
            excel_3p_monthly[m][plat] += order["subtotal"]
            excel_3p_counts[m][plat] += 1

    # ── 4. Notion: non-Toast 3P (primary for months it has data) ──
    notion_orders = _fetch_notion_3p()
    notion_3p_monthly = defaultdict(lambda: defaultdict(float))
    notion_3p_counts = defaultdict(lambda: defaultdict(int))
    notion_months = set()

    for o in notion_orders:
        m = o["date"][:7]
        notion_months.add(m)
        plat = o["platform"]
        notion_3p_monthly[m][plat] += o["subtotal"]
        notion_3p_counts[m][plat] += 1

    # ── 5. Washington Square CSV (Toast gap-fill + customer enrichment) ──
    ws_orders = _parse_ws_csv()

    # Aggregate WS CSV by month for Toast gap-filling
    ws_toast_monthly = defaultdict(float)
    ws_toast_counts_monthly = defaultdict(int)
    for o in ws_orders:
        m = o["date"][:7]
        ws_toast_monthly[m] += o.get("subtotal", 0)
        ws_toast_counts_monthly[m] += 1

    # ── 6. Merge all sources ──
    all_3p_platforms = {"Forkable", "EZCater", "Cater2me"}
    all_months = set(toast_revenue_monthly.keys())
    all_months.update(ws_toast_monthly.keys())
    for m in excel_3p_monthly:
        all_months.add(m)
    for m in notion_months:
        all_months.add(m)

    merged_monthly = defaultdict(lambda: defaultdict(float))
    merged_counts = defaultdict(lambda: defaultdict(int))

    for m in all_months:
        # Toast: max(POS, WS CSV) per month (same orders, partial POS cache)
        toast_pos = toast_revenue_monthly.get(m, 0)
        toast_ws = ws_toast_monthly.get(m, 0)
        toast_rev = max(toast_pos, toast_ws)
        if toast_rev > 0:
            merged_monthly[m]["Toast"] = round(toast_rev, 2)
            # Use whichever source had the higher revenue for counts too
            if toast_ws >= toast_pos:
                merged_counts[m]["Toast"] = ws_toast_counts_monthly.get(m, 1)
            else:
                merged_counts[m]["Toast"] = toast_order_counts_monthly.get(m, 1)

        # 3P: use Notion if it has data for this month, else Excel
        if m in notion_months:
            for plat in all_3p_platforms:
                val = notion_3p_monthly[m].get(plat, 0)
                if val > 0:
                    merged_monthly[m][plat] = round(val, 2)
                    merged_counts[m][plat] = notion_3p_counts[m].get(plat, 0)
        else:
            for plat in all_3p_platforms:
                val = excel_3p_monthly[m].get(plat, 0)
                if val > 0:
                    merged_monthly[m][plat] = round(val, 2)
                    merged_counts[m][plat] = excel_3p_counts[m].get(plat, 0)

    # ── 7. Apply period filter ──
    if start_month or end_month:
        filtered_months = set()
        for m in all_months:
            if start_month and m < start_month:
                continue
            if end_month and m > end_month:
                continue
            filtered_months.add(m)
        all_months = filtered_months
        ws_orders = [
            o for o in ws_orders
            if (not start_month or o["date"][:7] >= start_month) and
               (not end_month or o["date"][:7] <= end_month)
        ]
        top_items_filtered = []
        # Re-scan items within period if filtered
        if start_month or end_month:
            item_agg = defaultdict(lambda: {"qty": 0, "revenue": 0.0})
            for folder in sorted(os.listdir(_TMP_DIR)) if os.path.isdir(_TMP_DIR) else []:
                if not (folder.isdigit() and len(folder) == 8):
                    continue
                ds = folder[:4] + "-" + folder[4:6] + "-" + folder[6:8]
                dm = ds[:7]
                if dm not in all_months:
                    continue
                item_file = os.path.join(_TMP_DIR, folder, "ItemSelectionDetails.csv")
                if not os.path.exists(item_file):
                    continue
                try:
                    df = pd.read_csv(item_file)
                    if "Menu" in df.columns and "Net Price" in df.columns:
                        cat = df[df["Menu"].astype(str).str.contains("Catering", case=False, na=False)]
                        name_col = "Menu Item" if "Menu Item" in df.columns else None
                        qty_col = "Qty" if "Qty" in df.columns else None
                        for _, row in cat.iterrows():
                            name = str(row[name_col]) if name_col else "Unknown"
                            qty = int(row[qty_col]) if qty_col and pd.notna(row[qty_col]) else 1
                            price = float(row["Net Price"]) if pd.notna(row["Net Price"]) else 0
                            item_agg[name]["qty"] += qty
                            item_agg[name]["revenue"] += price
                except Exception as e:
                    logger.warning("Catering period items parse failed for %s: %s", folder, e)
            top_items = [
                {"item": n, "qty": d["qty"], "revenue": round(d["revenue"], 2)}
                for n, d in item_agg.items()
            ]
            top_items.sort(key=lambda x: x["revenue"], reverse=True)
            top_items = top_items[:20]

    sorted_months = sorted(all_months)

    if not sorted_months:
        return {
            "period": {"start": start_month or "", "end": end_month or "",
                       "months": 0, "is_filtered": bool(start_month or end_month),
                       "_start_month": start_month, "_end_month": end_month},
            "monthly": {"months": [], "platforms": {}, "totals": [],
                        "order_counts": {}, "fees": {}},
            "kpis": {},
            "top_items": [],
            "customers": {"top_companies": [], "top_customers": [], "repeat_rate": 0},
            "recent_orders": [],
            "dow_analysis": {"labels": [], "order_counts": [], "avg_revenue": []},
            "order_sources": {"labels": [], "counts": []},
        }

    # ── 8. Build monthly output ──
    platform_order = ["Toast", "Forkable", "EZCater", "Cater2me"]
    all_plats = set()
    for m in sorted_months:
        all_plats.update(merged_monthly[m].keys())
    ordered_plats = [p for p in platform_order if p in all_plats]
    for p in sorted(all_plats):
        if p not in ordered_plats:
            ordered_plats.append(p)

    monthly_out = {
        "months": sorted_months,
        "platforms": {},
        "totals": [],
        "order_counts": {},
        "fees": {m: round(fees_monthly.get(m, 0), 2) for m in sorted_months},
    }

    for plat in ordered_plats:
        monthly_out["platforms"][plat] = [
            round(merged_monthly[m].get(plat, 0), 2) for m in sorted_months
        ]
        monthly_out["order_counts"][plat] = [
            merged_counts[m].get(plat, 0) for m in sorted_months
        ]

    monthly_out["totals"] = [
        round(sum(merged_monthly[m].values()), 2) for m in sorted_months
    ]

    # ── 9. KPIs ──
    total_revenue = sum(monthly_out["totals"])
    toast_total = sum(
        merged_monthly[m].get("Toast", 0) for m in sorted_months
    )
    third_party_total = total_revenue - toast_total
    total_orders = sum(
        sum(merged_counts[m].values()) for m in sorted_months
    )
    avg_order = total_revenue / total_orders if total_orders else 0
    total_fees = sum(fees_monthly.get(m, 0) for m in sorted_months)
    total_tips = sum(o.get("tip", 0) for o in ws_orders)

    kpis = {
        "total_revenue": round(total_revenue, 2),
        "toast_revenue": round(toast_total, 2),
        "third_party_revenue": round(third_party_total, 2),
        "total_orders": total_orders,
        "avg_order_size": round(avg_order, 2),
        "total_fees": round(total_fees, 2),
        "total_tips": round(total_tips, 2),
    }

    # ── 10. Customer analysis (from WS CSV) ──
    company_agg = defaultdict(lambda: {"orders": 0, "total": 0.0})
    customer_agg = defaultdict(lambda: {"orders": 0, "total": 0.0, "company": ""})

    for o in ws_orders:
        company = o.get("company") or ""
        customer = o.get("customer") or ""
        sub = o.get("subtotal", 0)

        if company:
            company_agg[company]["orders"] += 1
            company_agg[company]["total"] += sub

        if customer:
            customer_agg[customer]["orders"] += 1
            customer_agg[customer]["total"] += sub
            if company and not customer_agg[customer]["company"]:
                customer_agg[customer]["company"] = company

    top_companies = [
        {
            "company": name,
            "orders": d["orders"],
            "total_spent": round(d["total"], 2),
            "avg_order": round(d["total"] / d["orders"], 2) if d["orders"] else 0,
        }
        for name, d in company_agg.items()
        if name
    ]
    top_companies.sort(key=lambda x: x["total_spent"], reverse=True)

    top_customers = [
        {
            "name": name,
            "company": d["company"],
            "orders": d["orders"],
            "total_spent": round(d["total"], 2),
        }
        for name, d in customer_agg.items()
        if name
    ]
    top_customers.sort(key=lambda x: x["total_spent"], reverse=True)

    repeat_customers = sum(1 for c in customer_agg.values() if c["orders"] > 1)
    total_unique = len(customer_agg)
    repeat_rate = (repeat_customers / total_unique * 100) if total_unique else 0

    customers = {
        "top_companies": top_companies[:15],
        "top_customers": top_customers[:15],
        "repeat_rate": round(repeat_rate, 1),
    }

    # ── 11. Recent orders (from WS CSV, sorted newest first) ──
    recent = sorted(ws_orders, key=lambda x: x["date"], reverse=True)[:25]
    recent_orders = [
        {
            "date": o["date"],
            "name": o.get("event_name") or o.get("customer") or "",
            "platform": o.get("platform", "Toast"),
            "subtotal": o.get("subtotal", 0),
            "company": o.get("company", ""),
            "customer": o.get("customer", ""),
            "guest_count": o.get("guest_count", 0),
            "city": o.get("city", ""),
        }
        for o in recent
    ]

    # Also include recent Notion 3P orders
    notion_recent = sorted(notion_orders, key=lambda x: x.get("date", ""), reverse=True)
    for o in notion_recent[:10]:
        m = o["date"][:7]
        if start_month and m < start_month:
            continue
        if end_month and m > end_month:
            continue
        recent_orders.append({
            "date": o["date"],
            "name": o.get("name", ""),
            "platform": o.get("platform", ""),
            "subtotal": o.get("subtotal", 0),
            "company": "",
            "customer": "",
            "guest_count": 0,
            "city": "",
        })

    recent_orders.sort(key=lambda x: x["date"], reverse=True)
    recent_orders = recent_orders[:25]

    # ── 12. All orders for click-to-detail modal ──
    all_orders = []

    # Toast orders from WS CSV (richest detail)
    for o in ws_orders:
        all_orders.append({
            "date": o["date"],
            "name": o.get("event_name") or o.get("customer") or "",
            "company": o.get("company", ""),
            "platform": "Toast",
            "subtotal": o.get("subtotal", 0),
            "guest_count": o.get("guest_count", 0),
            "city": o.get("city", ""),
            "has_detail": True,
        })

    # 3P: Notion months first, Excel fallback for remaining
    notion_3p_months = set()
    for o in notion_orders:
        m = o["date"][:7]
        if start_month and m < start_month:
            continue
        if end_month and m > end_month:
            continue
        notion_3p_months.add(m)
        all_orders.append({
            "date": o["date"],
            "name": o.get("name", ""),
            "company": "",
            "platform": o.get("platform", ""),
            "subtotal": o.get("subtotal", 0),
            "guest_count": 0,
            "city": "",
            "has_detail": True,
        })

    for o in tracker_orders:
        if o["platform"] == "_fees":
            continue
        m = o["date"][:7]
        if start_month and m < start_month:
            continue
        if end_month and m > end_month:
            continue
        if m in notion_3p_months:
            continue
        all_orders.append({
            "date": o["date"],
            "name": "",
            "company": "",
            "platform": o["platform"],
            "subtotal": o.get("subtotal", 0),
            "guest_count": 0,
            "city": "",
            "has_detail": False,
        })

    # ── 14. Day-of-week analysis (from WS CSV) ──
    dow_names = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    dow_rev = defaultdict(list)
    dow_count = defaultdict(int)

    for o in ws_orders:
        try:
            dt = datetime.strptime(o["date"], "%Y-%m-%d")
            dow = dt.weekday()  # 0=Mon
            dow_rev[dow].append(o.get("subtotal", 0))
            dow_count[dow] += 1
        except ValueError:
            pass

    dow_analysis = {
        "labels": dow_names,
        "order_counts": [dow_count.get(i, 0) for i in range(7)],
        "avg_revenue": [
            round(sum(dow_rev[i]) / len(dow_rev[i]), 2) if dow_rev[i] else 0
            for i in range(7)
        ],
    }

    # ── 15. Order sources (from WS CSV) ──
    source_counts = defaultdict(int)
    for o in ws_orders:
        src = o.get("created_from", "Unknown") or "Unknown"
        source_counts[src] += 1

    source_sorted = sorted(source_counts.items(), key=lambda x: x[1], reverse=True)
    order_sources = {
        "labels": [s[0] for s in source_sorted],
        "counts": [s[1] for s in source_sorted],
    }

    # ── 16. Period info ──
    period = {
        "start": sorted_months[0] if sorted_months else "",
        "end": sorted_months[-1] if sorted_months else "",
        "months": len(sorted_months),
        "is_filtered": bool(start_month or end_month),
        "_start_month": start_month,
        "_end_month": end_month,
    }

    result = {
        "period": period,
        "monthly": monthly_out,
        "kpis": kpis,
        "top_items": top_items,
        "customers": customers,
        "recent_orders": recent_orders,
        "all_orders": all_orders,
        "dow_analysis": dow_analysis,
        "order_sources": order_sources,
    }

    _dash_cache["ts"] = now
    _dash_cache["data"] = result
    return result


def get_catering_by_month():
    """Get catering revenue aggregated by month and platform.

    Combines:
      - Excel Tracker: Forkable, EZCater, Cater2Me, Catering Delivery
      - Toast POS cache: Catering Delivery dining option (supplements Excel)

    For Toast Catering, takes max(Excel, POS) per month to avoid double-counting.

    Returns dict with keys:
        months: list of "YYYY-MM" strings
        platforms: {platform_name: [monthly_subtotals]}
        totals: [monthly_totals]
        order_counts: {platform_name: [monthly_day_counts]}
        fees: {month: total_fees}
    Returns None if no data available.
    """
    now = time.time()
    if _cache["data"] is not None and (now - _cache["ts"]) < _CACHE_TTL:
        return _cache["data"]

    # Parse Excel tracker
    tracker_orders, excel_toast = _parse_tracker_catering()

    # Scan Toast POS cache
    pos_toast = _scan_toast_catering()

    # Merge Toast Catering: take max per month from either source
    all_toast_months = set(excel_toast.keys()) | set(pos_toast.keys())
    toast_catering = {}
    for m in all_toast_months:
        toast_catering[m] = max(excel_toast.get(m, 0), pos_toast.get(m, 0))

    # Aggregate 3P platforms by month from tracker orders
    monthly = defaultdict(lambda: defaultdict(float))
    counts = defaultdict(lambda: defaultdict(int))
    fees = defaultdict(float)

    for order in tracker_orders:
        month = order["date"][:7]
        platform = order["platform"]

        if platform == "_fees":
            fees[month] += order["subtotal"]
            continue

        monthly[month][platform] += order["subtotal"]
        counts[month][platform] += 1

    # Add Toast Catering
    for m, total in toast_catering.items():
        if total > 0:
            monthly[m]["Toast Catering"] = round(total, 2)
            counts[m]["Toast Catering"] = 1  # day count not meaningful here

    if not monthly:
        return None

    sorted_months = sorted(monthly.keys())

    # Collect and order platforms
    all_platforms = set()
    for m in sorted_months:
        all_platforms.update(monthly[m].keys())

    preferred_order = ["Forkable", "EZCater", "Cater2me", "Toast Catering"]
    ordered_platforms = [p for p in preferred_order if p in all_platforms]
    for p in sorted(all_platforms):
        if p not in ordered_platforms:
            ordered_platforms.append(p)

    result = {
        "months": sorted_months,
        "platforms": {},
        "totals": [],
        "order_counts": {},
        "fees": dict(fees),
    }

    for platform in ordered_platforms:
        result["platforms"][platform] = [
            round(monthly[m].get(platform, 0), 2) for m in sorted_months
        ]
        result["order_counts"][platform] = [
            counts[m].get(platform, 0) for m in sorted_months
        ]

    result["totals"] = [
        round(sum(monthly[m].values()), 2) for m in sorted_months
    ]

    _cache["ts"] = now
    _cache["data"] = result
    return result
