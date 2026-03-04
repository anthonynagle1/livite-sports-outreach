"""Backfill historical profit data from xlsx to Notion.

Reads the Tracker 2 spreadsheet and upserts each row into the Notion Profit Tracker DB.
Rate-limited to avoid hitting Notion API limits.

Usage:
    python tools/backfill_notion_profit.py
"""
import os
import sys
import time
from datetime import datetime

from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv(os.path.join(os.path.dirname(__file__), '..', '.env'))

sys.path.insert(0, os.path.dirname(__file__))
from profit_notion import upsert_daily_profit

NOTION_PROFIT_DB_ID = os.getenv("NOTION_PROFIT_DB_ID", "")
XLSX_PATH = os.path.join(os.path.dirname(__file__), '..', 'data', 'New Profit Calc (1).xlsx')

# Column mapping (1-indexed from the xlsx)
COL_MAP = {
    1: "date_raw",       # Date
    2: "profit",         # Profit
    3: "total_sales",    # Total Sales
    4: "toast_total",    # Toast Total
    5: "labor",          # Labor
    6: "vacation_sick",  # Vacation/Sick Time
    7: "misc",           # Misc
    8: "fixed",          # Fixed
    9: "service_fees",   # Service Fees
    10: "food_cost",     # Food Cost
    11: "forkable",      # Forkable
    12: "ezcater",       # EZ Cater
    13: "cater2me",      # Cater2Me
    27: "ot_hours",      # OT Hours
    28: "ot_pay",        # OT Pay
    29: "total_hours",   # Hourly Hours (maps to total_hours minus zadith)
    30: "blended_rate",  # Avg Wage
    31: "mgmt_hours",    # Mgmt Hours
    32: "zadith_hours",  # Zadith Hours
    33: "laurie_h",      # Laurie H
    34: "anthony_h",     # Anthony H
    35: "steve_h",       # Steve H
    36: "catering_fees", # Catering Fees
    37: "tds_fees",      # TDS Fees
    38: "gh_fees",       # GH Fees
    39: "dd_fees",       # DD Fees
    40: "uber_fees",     # Regular Uber Fees
    41: "uber_ads",      # Uber Ads
    45: "weather",       # Weather
    46: "notes",         # Notes
    47: "profit_pct",    # % Profit
}


def _num(val):
    """Convert cell value to float, None/empty -> 0."""
    if val is None:
        return 0
    try:
        f = float(val)
        return f if f == f else 0  # NaN check
    except (TypeError, ValueError):
        return 0


def _str(val):
    """Convert cell value to string."""
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s in ("N/A", "None", "nan") else s


def _upsert_with_retry(db_id, data, max_retries=3):
    """Upsert with retry on timeout/network errors."""
    for attempt in range(max_retries):
        try:
            return upsert_daily_profit(db_id, data)
        except Exception as e:
            if attempt < max_retries - 1:
                wait = 5 * (attempt + 1)
                print(f"    Retry {attempt+1}/{max_retries} after error: {e} (waiting {wait}s)")
                time.sleep(wait)
            else:
                print(f"    FAILED after {max_retries} retries: {e}")
                return None


def main():
    if not NOTION_PROFIT_DB_ID:
        print("ERROR: NOTION_PROFIT_DB_ID not set. Add it to .env")
        sys.exit(1)

    if not os.path.exists(XLSX_PATH):
        print(f"ERROR: File not found: {XLSX_PATH}")
        sys.exit(1)

    # Optional: start from a specific date (YYYYMMDD) to resume
    start_from = None
    if len(sys.argv) > 1:
        start_from = datetime.strptime(sys.argv[1], "%Y%m%d")
        print(f"Resuming from {start_from.strftime('%m/%d/%Y')}...")

    print(f"Loading {XLSX_PATH}...")
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb['Tracker']

    rows_processed = 0
    rows_success = 0
    rows_skipped = 0

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Extract date (col 1)
        if not row or len(row) == 0:
            continue
        date_val = row[0]
        if date_val is None:
            continue

        # Parse date
        if isinstance(date_val, datetime):
            dt = date_val
        elif isinstance(date_val, str):
            try:
                dt = datetime.strptime(date_val, "%Y-%m-%d")
            except ValueError:
                try:
                    dt = datetime.strptime(date_val, "%m/%d/%Y")
                except ValueError:
                    print(f"  Row {row_num}: Skipping, can't parse date: {date_val}")
                    rows_skipped += 1
                    continue
        else:
            print(f"  Row {row_num}: Skipping, unexpected date type: {type(date_val)}")
            rows_skipped += 1
            continue

        # Skip rows before start_from date
        if start_from and dt < start_from:
            continue

        date_str = dt.strftime("%Y%m%d")
        display_date = dt.strftime("%m/%d/%Y")
        day_of_week = dt.strftime("%A")

        # Build data dict
        data = {
            "date": date_str,
            "display_date": display_date,
            "day_of_week": day_of_week,
        }

        # Extract numeric fields
        for col_idx, key in COL_MAP.items():
            if col_idx == 1:
                continue  # date handled above
            val = row[col_idx - 1] if col_idx - 1 < len(row) else None
            if key in ("weather", "notes"):
                data[key] = _str(val)
            elif key == "profit_pct":
                # Stored as decimal in xlsx (0.295 = 29.5%)
                pct = _num(val)
                data[key] = round(pct * 100, 2) if pct < 1 else round(pct, 2)
            else:
                data[key] = _num(val)

        # Compute total_hours = hourly hours + zadith
        hourly_hours = _num(row[28])  # col 29 (0-indexed: 28)
        zadith_h = data.get("zadith_hours", 0)
        data["total_hours"] = round(hourly_hours + zadith_h, 2)

        # FTEs
        th = data.get("total_hours", 0)
        data["ftes"] = round(th / 8, 2) if th > 0 else 0

        # Payroll taxes (not in xlsx directly, estimate: total_hours * $3)
        data["payroll_taxes"] = round(th * 3, 2)

        # Put manual fields under "manual" key for _build_properties compatibility
        manual_keys = ["forkable", "ezcater", "cater2me", "fixed", "vacation_sick",
                       "misc", "gh_fees", "dd_fees", "uber_fees", "uber_ads",
                       "catering_fees", "mgmt_hours",
                       "laurie_h", "anthony_h", "steve_h", "weather", "notes"]
        manual = {}
        for k in manual_keys:
            if k in data and data[k]:
                manual[k] = data[k]
        data["manual"] = manual

        rows_processed += 1

        # Upsert to Notion with retry
        result = _upsert_with_retry(NOTION_PROFIT_DB_ID, data)
        if result:
            rows_success += 1
            print(f"  {display_date} ({day_of_week}): OK — Profit ${data.get('profit', 0):,.0f}")
        else:
            print(f"  {display_date}: FAILED")

        # Rate limit: ~3 requests/sec (Notion limit is ~3/sec)
        time.sleep(0.4)

        # Progress update every 50 rows
        if rows_processed % 50 == 0:
            print(f"\n  --- Progress: {rows_processed} processed, {rows_success} success, {rows_skipped} skipped ---\n")

    wb.close()

    print(f"\n{'='*50}")
    print(f"  BACKFILL COMPLETE")
    print(f"  Processed: {rows_processed}")
    print(f"  Success:   {rows_success}")
    print(f"  Skipped:   {rows_skipped}")
    print(f"  Failed:    {rows_processed - rows_success}")
    print(f"{'='*50}")


if __name__ == "__main__":
    main()
