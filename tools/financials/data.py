"""
Parse P&L and Balance Sheet Excel files into unified, chronologically sorted dicts.

Data source: Accounting system exports (Bench/similar) for Square Press LLC - Livite.
Covers Nov 2016 - Jan 2026 across multiple Excel files.
"""

from __future__ import annotations

import os
from collections import defaultdict
from datetime import datetime

import openpyxl

_BASE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# ---------------------------------------------------------------------------
# File configuration
# ---------------------------------------------------------------------------

PL_FILES = [
    os.path.join(_BASE, "data", "financials", "Square_Press_LLC_-_Livite_-_Income_Statement__Profit_and_Loss_ (9).xlsx"),
    os.path.join(_BASE, "data", "financials", "Square_Press_LLC_-_Livite_-_Income_Statement__Profit_and_Loss_ (8).xlsx"),
    os.path.join(_BASE, "data", "financials", "Square_Press_LLC_-_Livite_-_Income_Statement__Profit_and_Loss_ (6).xlsx"),
]

BS_FILES = [
    os.path.join(_BASE, "data", "financials", "Square_Press_LLC_-_Livite_-_Balance_Sheet (7).xlsx"),
    os.path.join(_BASE, "data", "financials", "Square_Press_LLC_-_Livite_-_Balance_Sheet (6).xlsx"),
    os.path.join(_BASE, "data", "financials", "Square_Press_LLC_-_Livite_-_Balance_Sheet (5).xlsx"),
]

# Section headers and total/formula row names to skip as data
_SECTION_HEADERS = {
    "Income", "Cost of Goods Sold", "Operating Expenses",
    "Other Income / (Expense)",
}

_TOTAL_ROW_PREFIXES = ("Total ", "Gross Profit", "Operating Income", "Net Income")

# ---------------------------------------------------------------------------
# Caching
# ---------------------------------------------------------------------------

_pl_cache: dict | None = None
_bs_cache: dict | None = None


def _parse_month_key(val: str) -> str | None:
    """Parse 'MMM YYYY' into '2024-01' format."""
    if not val or not isinstance(val, str):
        return None
    val = val.strip()
    try:
        dt = datetime.strptime(val, "%b %Y")
        return dt.strftime("%Y-%m")
    except ValueError:
        return None


def _parse_bs_date(val: str) -> str | None:
    """Parse 'MMM DD, YYYY' into '2024-01' format."""
    if not val or not isinstance(val, str):
        return None
    val = val.strip()
    try:
        dt = datetime.strptime(val, "%b %d, %Y")
        return dt.strftime("%Y-%m")
    except ValueError:
        return None


def _is_section_header(row_cells: list, data_start_col: int) -> bool:
    """Section headers have None in all data columns."""
    for cell in row_cells[data_start_col:]:
        if cell.value is not None:
            return False
    return True


def _is_total_row(name: str) -> bool:
    """Check if row is a formula/total row that should be recalculated."""
    if not name:
        return False
    for prefix in _TOTAL_ROW_PREFIXES:
        if name.startswith(prefix):
            return True
    return False


# ---------------------------------------------------------------------------
# P&L Parser
# ---------------------------------------------------------------------------

def _parse_single_pl(filepath: str) -> dict:
    """Parse a single P&L Excel file.

    Returns:
        {
            "months": [list of 'YYYY-MM' keys, reverse-chron],
            "sections": {
                "income": {"Account Name": {"YYYY-MM": value, ...}, ...},
                "cogs": {"Cost of Goods Sold": {...}},
                "opex": {"Account Name": {...}, ...},
                "other_income": {"Account Name": {...}, ...},
            }
        }
    """
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=1, values_only=False))
    wb.close()

    # Row 5 (index 4) has date headers in columns B onward
    header_row = rows[4]
    months = []
    col_to_month = {}
    for i, cell in enumerate(header_row):
        if i == 0:
            continue  # Column A = "Account"
        mk = _parse_month_key(cell.value)
        if mk:
            months.append(mk)
            col_to_month[i] = mk

    # Track current section
    current_section = None
    sections = {
        "income": {},
        "cogs": {},
        "opex": {},
        "other_income": {},
    }

    section_map = {
        "Income": "income",
        "Cost of Goods Sold": "cogs",
        "Operating Expenses": "opex",
        "Other Income / (Expense)": "other_income",
    }

    for row in rows[5:]:  # Start after header row
        name = row[0].value
        if not name or not isinstance(name, str):
            continue
        name = name.strip()

        # Check for section header
        if name in _SECTION_HEADERS and _is_section_header(row, 1):
            current_section = section_map.get(name)
            continue

        # Skip total/formula rows
        if _is_total_row(name):
            continue

        if current_section is None:
            continue

        # Read data values
        values = {}
        for col_idx, month_key in col_to_month.items():
            if col_idx < len(row):
                val = row[col_idx].value
                if val is not None:
                    try:
                        values[month_key] = float(val)
                    except (TypeError, ValueError):
                        values[month_key] = 0.0
                else:
                    values[month_key] = 0.0
            else:
                values[month_key] = 0.0

        sections[current_section][name] = values

    return {"months": months, "sections": sections}


def parse_all_pl() -> dict:
    """Parse all P&L files and merge into one unified timeline.

    Returns:
        {
            "months": ["2016-11", "2016-12", ..., "2026-01"],
            "income": {"Sales": [val, val, ...], "Discounts": [...], ...},
            "cogs": [val, val, ...],
            "opex": {"Account": [val, val, ...], ...},
            "other_income": {"Account": [val, val, ...], ...},
            "total_income": [val, ...],
            "gross_profit": [val, ...],
            "total_opex": [val, ...],
            "operating_income": [val, ...],
            "total_other_income": [val, ...],
            "net_income": [val, ...],
        }
    """
    global _pl_cache
    if _pl_cache is not None:
        return _pl_cache

    # Collect all data keyed by month
    all_income: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    all_cogs: dict[str, float] = defaultdict(float)
    all_opex: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    all_other: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    all_months: set[str] = set()

    for filepath in PL_FILES:
        if not os.path.exists(filepath):
            continue
        parsed = _parse_single_pl(filepath)
        all_months.update(parsed["months"])

        for acct, vals in parsed["sections"]["income"].items():
            for mk, v in vals.items():
                all_income[acct][mk] = v

        for acct, vals in parsed["sections"]["cogs"].items():
            for mk, v in vals.items():
                all_cogs[mk] += v  # Combine if multiple COGS lines

        for acct, vals in parsed["sections"]["opex"].items():
            for mk, v in vals.items():
                all_opex[acct][mk] = v

        for acct, vals in parsed["sections"]["other_income"].items():
            for mk, v in vals.items():
                all_other[acct][mk] = v

    # Sort months chronologically
    months = sorted(all_months)

    # Build arrays aligned to months
    income = {}
    for acct in sorted(all_income.keys()):
        income[acct] = [all_income[acct].get(m, 0.0) for m in months]

    cogs = [all_cogs.get(m, 0.0) for m in months]

    opex = {}
    for acct in sorted(all_opex.keys()):
        opex[acct] = [all_opex[acct].get(m, 0.0) for m in months]

    other_income = {}
    for acct in sorted(all_other.keys()):
        other_income[acct] = [all_other[acct].get(m, 0.0) for m in months]

    # Recalculate totals
    n = len(months)
    total_income = [0.0] * n
    for vals in income.values():
        for i in range(n):
            total_income[i] += vals[i]

    gross_profit = [total_income[i] - cogs[i] for i in range(n)]

    total_opex = [0.0] * n
    for vals in opex.values():
        for i in range(n):
            total_opex[i] += vals[i]

    operating_income = [gross_profit[i] - total_opex[i] for i in range(n)]

    total_other_income = [0.0] * n
    for vals in other_income.values():
        for i in range(n):
            total_other_income[i] += vals[i]

    net_income = [operating_income[i] + total_other_income[i] for i in range(n)]

    result = {
        "months": months,
        "income": income,
        "cogs": cogs,
        "opex": opex,
        "other_income": other_income,
        "total_income": total_income,
        "gross_profit": gross_profit,
        "total_opex": total_opex,
        "operating_income": operating_income,
        "total_other_income": total_other_income,
        "net_income": net_income,
    }
    _pl_cache = result
    return result


# ---------------------------------------------------------------------------
# Balance Sheet Parser
# ---------------------------------------------------------------------------

def _parse_single_bs(filepath: str) -> dict:
    """Parse a single Balance Sheet Excel file.

    Returns:
        {
            "months": [list of 'YYYY-MM' keys],
            "accounts": {"Account Name": {"YYYY-MM": value, ...}, ...},
            "sections": {"Account Name": "section_key"},  # e.g. "current_assets"
        }
    """
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=1, values_only=False))
    wb.close()

    # Find date headers - could be in row 5, columns D onward
    # BS files have Account spanning cols A-C, data from col D
    header_row = rows[4]  # row 5
    months = []
    col_to_month = {}

    for i, cell in enumerate(header_row):
        if i < 3:
            continue  # Skip A, B, C (account name columns)
        mk = _parse_bs_date(cell.value)
        if mk:
            months.append(mk)
            col_to_month[i] = mk

    accounts: dict[str, dict[str, float]] = {}
    account_sections: dict[str, str] = {}
    current_section = None

    # Section tracking from column A headers
    section_keywords = {
        "Assets": "assets",
        "Current Assets": "current_assets",
        "Cash and Cash Equivalents": "cash",
        "Fixed Assets": "fixed_assets",
        "Long Term Assets": "long_term_assets",
        "Liabilities and Equity": "liabilities_equity",
        "Current Liabilities": "current_liabilities",
        "Long Term Liabilities": "long_term_liabilities",
        "Equity": "equity",
    }

    for row in rows[5:]:
        # Get account name from first non-None of A, B, C
        name = None
        depth = 0
        for col_idx in range(min(3, len(row))):
            if row[col_idx].value is not None:
                name = str(row[col_idx].value).strip()
                depth = col_idx
                break

        if not name:
            continue

        # Check for section header
        if name in section_keywords:
            current_section = section_keywords[name]
            continue

        # Skip total rows
        if _is_total_row(name):
            continue

        # Skip empty data rows (section headers at deeper levels)
        has_data = False
        for col_idx in col_to_month:
            if col_idx < len(row) and row[col_idx].value is not None:
                has_data = True
                break
        if not has_data:
            continue

        # Read values
        values = {}
        for col_idx, month_key in col_to_month.items():
            if col_idx < len(row):
                val = row[col_idx].value
                if val is not None:
                    try:
                        values[month_key] = float(val)
                    except (TypeError, ValueError):
                        values[month_key] = 0.0
                else:
                    values[month_key] = 0.0
            else:
                values[month_key] = 0.0

        accounts[name] = values
        if current_section:
            account_sections[name] = current_section

    return {"months": months, "accounts": accounts, "sections": account_sections}


# Account classification for balance sheet grouping
_CASH_ACCOUNTS = {
    "BOA Business Account 3994", "Brookline Checking 0270", "TD Bank",
    "American Express 1009/2007", "American Express 31004",
    "American Express 610000", "BOA CC 1032", "BOA CC 4040",
    "First Bankcard CC 1064", "American Express 1111",
}

_CURRENT_ASSET_OTHER = {
    "Cash Clearing", "Credit Card Clearing", "Inventory",
    "Payroll Clearing", "Payroll Reconciliation Refund Receivable",
    "Deposits in Transit", "ERC Receivable", "Transfer Clearing",
}

_FIXED_ASSET_ACCOUNTS = {
    "Accumulated Amortization", "Accumulated Depreciation",
    "Fixed Assets", "Intangible Assets",
}

_LONG_TERM_ASSET_ACCOUNTS = {
    "Investments", "Lease Security Deposit Receivable",
}

_LONG_TERM_LIABILITY_ACCOUNTS = {
    "Brookline Loan", "Steve & Laurie Catchup Distribution Liability",
}

_EQUITY_ACCOUNTS = {
    "Anthony Shareholder Distributions", "Capital Contributions",
    "Current Year Earnings", "Retained Earnings",
    "Steve & Laurie Shareholder Distributions",
}


def parse_all_bs() -> dict:
    """Parse all Balance Sheet files and merge into unified timeline.

    Returns:
        {
            "months": ["2017-01", ...],
            "cash": {"Account": [val, ...], ...},
            "total_cash": [val, ...],
            "current_assets_other": {"Account": [val, ...], ...},
            "total_current_assets": [val, ...],
            "fixed_assets": {"Account": [val, ...], ...},
            "total_fixed_assets": [val, ...],
            "long_term_assets": {"Account": [val, ...], ...},
            "total_long_term_assets": [val, ...],
            "total_assets": [val, ...],
            "current_liabilities": {"Account": [val, ...], ...},
            "total_current_liabilities": [val, ...],
            "long_term_liabilities": {"Account": [val, ...], ...},
            "total_long_term_liabilities": [val, ...],
            "total_liabilities": [val, ...],
            "equity": {"Account": [val, ...], ...},
            "total_equity": [val, ...],
        }
    """
    global _bs_cache
    if _bs_cache is not None:
        return _bs_cache

    # Collect all account data keyed by month
    all_accounts: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    all_months: set[str] = set()

    for filepath in BS_FILES:
        if not os.path.exists(filepath):
            continue
        parsed = _parse_single_bs(filepath)
        all_months.update(parsed["months"])
        for acct, vals in parsed["accounts"].items():
            for mk, v in vals.items():
                # Later files overwrite earlier for overlap months
                all_accounts[acct][mk] = v

    months = sorted(all_months)
    n = len(months)

    def _build_group(account_names: set) -> tuple[dict, list]:
        group = {}
        totals = [0.0] * n
        for acct in sorted(all_accounts.keys()):
            if acct in account_names:
                vals = [all_accounts[acct].get(m, 0.0) for m in months]
                group[acct] = vals
                for i in range(n):
                    totals[i] += vals[i]
        return group, totals

    # Classify accounts that aren't in known sets as current liabilities
    known = (_CASH_ACCOUNTS | _CURRENT_ASSET_OTHER | _FIXED_ASSET_ACCOUNTS |
             _LONG_TERM_ASSET_ACCOUNTS | _LONG_TERM_LIABILITY_ACCOUNTS | _EQUITY_ACCOUNTS)
    current_liab_accounts = set()
    for acct in all_accounts:
        if acct not in known:
            current_liab_accounts.add(acct)

    cash, total_cash = _build_group(_CASH_ACCOUNTS)
    ca_other, total_ca_other = _build_group(_CURRENT_ASSET_OTHER)
    total_current_assets = [total_cash[i] + total_ca_other[i] for i in range(n)]

    fixed, total_fixed = _build_group(_FIXED_ASSET_ACCOUNTS)
    lt_assets, total_lt_assets = _build_group(_LONG_TERM_ASSET_ACCOUNTS)
    total_assets = [total_current_assets[i] + total_fixed[i] + total_lt_assets[i] for i in range(n)]

    cur_liab, total_cur_liab = _build_group(current_liab_accounts)
    lt_liab, total_lt_liab = _build_group(_LONG_TERM_LIABILITY_ACCOUNTS)
    total_liabilities = [total_cur_liab[i] + total_lt_liab[i] for i in range(n)]

    equity, total_equity = _build_group(_EQUITY_ACCOUNTS)

    result = {
        "months": months,
        "cash": cash,
        "total_cash": total_cash,
        "current_assets_other": ca_other,
        "total_current_assets": total_current_assets,
        "fixed_assets": fixed,
        "total_fixed_assets": total_fixed,
        "long_term_assets": lt_assets,
        "total_long_term_assets": total_lt_assets,
        "total_assets": total_assets,
        "current_liabilities": cur_liab,
        "total_current_liabilities": total_cur_liab,
        "long_term_liabilities": lt_liab,
        "total_long_term_liabilities": total_lt_liab,
        "total_liabilities": total_liabilities,
        "equity": equity,
        "total_equity": total_equity,
    }
    _bs_cache = result
    return result
