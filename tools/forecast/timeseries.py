"""Build a daily sales time series from cached Toast POS data + Notion.

Scans .tmp/YYYYMMDD/OrderDetails.csv for all cached dates, aggregates
revenue by channel group (In-Store, Delivery, Catering), and joins
weather data from the Excel Tracker. 3P catering revenue (Forkable,
EZCater, Cater2me) is pulled from the Notion catering database.

This is Step 1 of the rolling P&L forecast engine.
"""

from __future__ import annotations

import logging
import os
import time
from collections import defaultdict
from datetime import datetime

import pandas as pd

logger = logging.getLogger(__name__)

# ── In-process DataFrame cache ───────────────────────────────────────────────
# Keeps the timeseries in memory so repeated requests (day-selector clicks,
# backtest runs, etc.) don't re-hit Azure Blob Storage every time.
_MEM_CACHE: dict = {"df": None, "ts": 0.0}
_MEM_TTL = 300  # seconds — refresh if >5 minutes old

_BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
_TMP_DIR = os.path.join(_BASE_DIR, ".tmp")
_CACHE_PATH = os.path.join(_TMP_DIR, "daily_timeseries.csv")
_EXCEL_PATH = os.path.join(_BASE_DIR, "data", "New Profit Calc (1).xlsx")

# ── Channel group classification ──
# Matches forecast doc: In-Store (~65%), Delivery (~25%), Catering (~10%)
_INSTORE = {
    "To Go",
    "Phone",
    "Online Ordering - Takeout",
}

_DELIVERY = {
    "Uber Eats - Delivery",
    "Uber Eats - Takeout",
    "DoorDash - Delivery",
    "DoorDash - Takeout",
    "Grubhub - Delivery",
    "Grubhub - Takeout",
    "Online Ordering - Delivery",
    "Delivery",
}

_CATERING = {
    "Catering Delivery",
}

_DOW_NAMES = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]


def _load_notion_3p_catering() -> dict:
    """Fetch all 3P catering orders from Notion, indexed by date.

    Returns dict: {YYYY-MM-DD: {"revenue": float, "orders": int}}
    Only includes non-Toast platforms (Forkable, EZCater, Cater2me).
    """
    try:
        from tools.catering.notion import _fetch_all_orders, NON_TOAST_PLATFORMS
    except ImportError:
        logger.warning("Could not import catering.notion — skipping 3P catering")
        return {}

    try:
        all_orders = _fetch_all_orders()
    except Exception as e:
        logger.warning("Notion 3P catering fetch failed: %s", e)
        return {}

    by_date = defaultdict(lambda: {"revenue": 0.0, "orders": 0})
    for order in all_orders:
        if order.get("platform") in NON_TOAST_PLATFORMS:
            date = order.get("date", "")
            if date:
                by_date[date]["revenue"] += order.get("subtotal", 0)
                by_date[date]["orders"] += 1

    logger.info("Loaded %d 3P catering dates from Notion", len(by_date))
    return dict(by_date)


def _classify_channel(dining_option: str) -> str:
    """Classify a Toast Dining Options value into a channel group."""
    if dining_option in _INSTORE:
        return "instore"
    if dining_option in _DELIVERY:
        return "delivery"
    if dining_option in _CATERING:
        return "catering"
    # E-Gift Cards and unknown go to "other"
    return "other"


def _load_weather() -> pd.DataFrame:
    """Load weather data from the Excel Tracker's Weather Data sheet."""
    if not os.path.exists(_EXCEL_PATH):
        logger.warning("Excel tracker not found: %s", _EXCEL_PATH)
        return pd.DataFrame()

    try:
        from openpyxl import load_workbook
        wb = load_workbook(_EXCEL_PATH, read_only=True, data_only=True)
    except Exception as e:
        logger.error("Failed to open Excel tracker: %s", e)
        return pd.DataFrame()

    if "Weather Data" not in wb.sheetnames:
        wb.close()
        return pd.DataFrame()

    ws = wb["Weather Data"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return pd.DataFrame()

    records = []
    for row in rows[1:]:
        if row[0] is None:
            continue
        if isinstance(row[0], datetime):
            date_str = row[0].strftime("%Y-%m-%d")
        else:
            continue

        def _num(idx):
            v = row[idx] if idx < len(row) else None
            if v is None:
                return None
            try:
                return float(v)
            except (ValueError, TypeError):
                return None

        records.append({
            "date": date_str,
            "temp_high": _num(1),
            "temp_low": _num(2),
            "precipitation_mm": _num(3),
            "snow_cm": _num(4),
            "wind_max_kmh": _num(5),
        })

    return pd.DataFrame(records)


def _scan_date_folders() -> list[str]:
    """Return sorted list of YYYYMMDD folders in .tmp/."""
    if not os.path.isdir(_TMP_DIR):
        return []
    folders = []
    for name in os.listdir(_TMP_DIR):
        if name.isdigit() and len(name) == 8:
            folders.append(name)
    return sorted(folders)


def _process_day(folder: str) -> dict | None:
    """Process one day's OrderDetails.csv into a summary row."""
    order_file = os.path.join(_TMP_DIR, folder, "OrderDetails.csv")
    if not os.path.exists(order_file):
        return None

    try:
        df = pd.read_csv(order_file)
    except Exception:
        return None

    if df.empty:
        return None

    # Filter voided orders
    if "Voided" in df.columns:
        df = df[df["Voided"] == False]

    if df.empty or "Amount" not in df.columns or "Dining Options" not in df.columns:
        return None

    date_str = folder[:4] + "-" + folder[4:6] + "-" + folder[6:8]

    # Classify and aggregate
    rev = defaultdict(float)
    disc = defaultdict(float)
    cnt = defaultdict(int)

    has_discount = "Discount Amount" in df.columns

    for _, row in df.iterrows():
        dining = str(row["Dining Options"]) if pd.notna(row["Dining Options"]) else ""
        group = _classify_channel(dining)
        amount = float(row["Amount"]) if pd.notna(row["Amount"]) else 0.0
        rev[group] += amount
        cnt[group] += 1
        if has_discount:
            d = row["Discount Amount"]
            disc[group] += abs(float(d)) if pd.notna(d) else 0.0

    total_rev = sum(rev.values())
    total_orders = sum(cnt.values())
    total_disc = sum(disc.values())

    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        dow = dt.weekday()
    except ValueError:
        dow = 0

    return {
        "date": date_str,
        "dow": dow,
        "dow_name": _DOW_NAMES[dow],
        "month": date_str[:7],
        "revenue_instore": round(rev.get("instore", 0), 2),
        "revenue_delivery": round(rev.get("delivery", 0), 2),
        "revenue_catering_toast": round(rev.get("catering", 0), 2),
        "revenue_other": round(rev.get("other", 0), 2),
        "discount_instore": round(disc.get("instore", 0), 2),
        "discount_delivery": round(disc.get("delivery", 0), 2),
        "discount_catering": round(disc.get("catering", 0), 2),
        "discount_total": round(total_disc, 2),
        "orders_instore": cnt.get("instore", 0),
        "orders_delivery": cnt.get("delivery", 0),
        "orders_catering_toast": cnt.get("catering", 0),
        "orders_total": total_orders,
    }


def _ensure_all_orderdetails():
    """Fetch OrderDetails.csv for ALL available dates from Azure Blob Storage.

    Only fetches dates that are missing from the local cache.
    This ensures the forecast always uses the complete history (~15 months),
    not just whatever dates happen to be cached from dashboard views.
    """
    try:
        from fetch_toast_data import (
            list_available_dates, get_toast_csv_cached,
        )
    except ImportError:
        logger.warning("fetch_toast_data not available — using cached data only")
        return

    try:
        all_dates = list_available_dates()
    except Exception as e:
        logger.warning("Could not list Azure dates: %s", e)
        return

    cached = set(_scan_date_folders())
    # Only need dates that don't have OrderDetails.csv cached
    missing = []
    for d in all_dates:
        order_file = os.path.join(_TMP_DIR, d, "OrderDetails.csv")
        if not os.path.exists(order_file):
            missing.append(d)

    if not missing:
        logger.info("All %d dates already cached", len(all_dates))
        return

    logger.info("Fetching OrderDetails.csv for %d uncached dates...", len(missing))
    for d in missing:
        try:
            get_toast_csv_cached(d, "OrderDetails.csv")
        except Exception as e:
            logger.debug("Failed to fetch %s: %s", d, e)


def build_daily_timeseries() -> pd.DataFrame:
    """Build the complete daily sales time series from Azure + .tmp/ cache.

    Auto-fetches OrderDetails.csv for ALL available dates from Azure Blob
    Storage before scanning. This ensures forecasts always use the full
    history (~15 months), not just whatever was previously cached.

    Returns pandas DataFrame with one row per day.
    """
    # Ensure ALL historical OrderDetails are cached locally
    _ensure_all_orderdetails()

    folders = _scan_date_folders()
    if not folders:
        logger.warning("No date folders found in %s", _TMP_DIR)
        return pd.DataFrame()

    logger.info("Building time series from %d date folders...", len(folders))

    # Pre-fetch all 3P catering from Notion (one API call, not per-day)
    notion_3p = _load_notion_3p_catering()

    rows = []
    for folder in folders:
        row = _process_day(folder)
        if row is not None:
            # Merge 3P catering from Notion
            date = row["date"]
            catering_3p = notion_3p.get(date, {"revenue": 0.0, "orders": 0})
            row["revenue_catering_3p"] = round(catering_3p["revenue"], 2)
            row["orders_catering_3p"] = catering_3p["orders"]

            # Combined catering and totals
            row["revenue_catering"] = round(
                row["revenue_catering_toast"] + row["revenue_catering_3p"], 2
            )
            row["orders_catering"] = (
                row["orders_catering_toast"] + row["orders_catering_3p"]
            )

            # Total revenue includes all channels + 3P catering
            toast_total = (
                row["revenue_instore"]
                + row["revenue_delivery"]
                + row["revenue_catering_toast"]
                + row["revenue_other"]
            )
            row["revenue_total"] = round(toast_total + row["revenue_catering_3p"], 2)
            row["orders_total"] = row["orders_total"] + row["orders_catering_3p"]
            row["avg_check"] = (
                round(row["revenue_total"] / row["orders_total"], 2)
                if row["orders_total"] else 0
            )

            rows.append(row)

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    df = df.sort_values("date").reset_index(drop=True)

    # Join weather data
    weather = _load_weather()
    if not weather.empty:
        df = df.merge(weather, on="date", how="left")
    else:
        df["temp_high"] = None
        df["temp_low"] = None
        df["precipitation_mm"] = None
        df["snow_cm"] = None
        df["wind_max_kmh"] = None

    # Write cache
    try:
        df.to_csv(_CACHE_PATH, index=False)
        logger.info("Wrote time series cache: %s (%d rows)", _CACHE_PATH, len(df))
    except Exception as e:
        logger.warning("Failed to write cache: %s", e)

    return df


def invalidate_timeseries_cache() -> None:
    """Force the in-memory timeseries cache to refresh on next call."""
    _MEM_CACHE["df"] = None
    _MEM_CACHE["ts"] = 0.0


def get_timeseries() -> pd.DataFrame:
    """Get the daily time series, using in-memory cache when fresh.

    Cache hierarchy:
    1. In-process memory cache (TTL: 5 min) — fastest, skips all I/O
    2. Disk CSV cache (.tmp/daily_timeseries.csv) — fast, skips Azure
    3. Full rebuild from Azure + .tmp/ folders — slow, only when needed

    Cache is invalidated when:
    - A date folder in .tmp/ is newer than the cache file, OR
    - The cache has significantly fewer rows than available date folders
      (meaning new dates were fetched from Azure since last build)
    """
    now = time.time()
    if _MEM_CACHE["df"] is not None and (now - _MEM_CACHE["ts"]) < _MEM_TTL:
        logger.debug("Timeseries: memory cache hit (%d rows)", len(_MEM_CACHE["df"]))
        return _MEM_CACHE["df"]

    if os.path.exists(_CACHE_PATH):
        cache_mtime = os.path.getmtime(_CACHE_PATH)

        # First ensure all Azure dates are cached locally
        _ensure_all_orderdetails()

        # Check if any date folder is newer than cache
        folders = _scan_date_folders()
        if folders:
            newest_folder = os.path.join(_TMP_DIR, folders[-1])
            folder_mtime = os.path.getmtime(newest_folder)
            if folder_mtime <= cache_mtime:
                try:
                    df = pd.read_csv(_CACHE_PATH)
                    # Invalidate if schema changed (new discount columns)
                    if "discount_total" not in df.columns:
                        logger.info("Cache missing discount columns — rebuilding")
                    # Invalidate if cache has far fewer rows than folders
                    elif len(df) >= len(folders) - 5:
                        logger.info("Loaded cached time series: %d rows", len(df))
                        _MEM_CACHE["df"] = df
                        _MEM_CACHE["ts"] = time.time()
                        return df
                    else:
                        logger.info(
                            "Cache stale: %d rows vs %d folders — rebuilding",
                            len(df), len(folders),
                        )
                except Exception:
                    pass

    df = build_daily_timeseries()
    if not df.empty:
        _MEM_CACHE["df"] = df
        _MEM_CACHE["ts"] = time.time()
    return df
